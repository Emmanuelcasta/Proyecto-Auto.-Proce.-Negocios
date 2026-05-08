"""
Servicio de reportes — Generación de Excel (PILA) y resúmenes.
"""

import io
import logging
from datetime import date, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.nomina import Nomina
from app.models.detalle_nomina import DetalleNomina
from app.utils.exceptions import SysClockException

logger = logging.getLogger("sysclock")

async def generar_excel_pila(db: AsyncSession, periodo: str) -> io.BytesIO:
    """
    Genera un archivo Excel con la información para la planilla PILA.
    periodo: "YYYY-MM"
    """
    try:
        year, month = map(int, periodo.split("-"))
        fecha_inicio = date(year, month, 1)
        if month == 12:
            fecha_fin = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            fecha_fin = date(year, month + 1, 1) - timedelta(days=1)
    except (ValueError, IndexError):
        raise SysClockException("Formato de periodo inválido. Use YYYY-MM.", status_code=400)

    # Buscar todas las nóminas aprobadas o pagadas del mes
    query = (
        select(Nomina)
        .where(
            Nomina.fecha_inicio >= fecha_inicio,
            Nomina.fecha_fin <= fecha_fin,
            Nomina.estado.in_(["APROBADO", "PAGADO"])
        )
        .options(selectinload(Nomina.empleado))
    )
    result = await db.execute(query)
    nominas = result.scalars().all()

    if not nominas:
        raise SysClockException(f"No hay nóminas aprobadas para el periodo {periodo}.", status_code=404)

    # Agrupar por empleado (sumar las dos quincenas si existen)
    datos_pila = {}
    for n in nominas:
        emp_id = n.empleado_id
        if emp_id not in datos_pila:
            datos_pila[emp_id] = {
                "nombre": n.empleado.nombre,
                "documento": n.empleado.cedula,
                "salario_base": n.empleado.salario,
                "ibc": Decimal("0"),
                "salud_emp": Decimal("0"),
                "pension_emp": Decimal("0"),
                "salud_pat": Decimal("0"),
                "pension_pat": Decimal("0"),
                "arl": Decimal("0"),
                "caja": Decimal("0"),
                "sena": Decimal("0"),
                "icbf": Decimal("0"),
            }
        
        # Obtener detalles para sumar aportes
        detalles_query = select(DetalleNomina).where(DetalleNomina.nomina_id == n.id)
        detalles_result = await db.execute(detalles_query)
        detalles = detalles_result.scalars().all()
        
        # IBC es la suma de salario básico proporcional (no incluye auxilio transporte)
        for d in detalles:
            if d.concepto == "Salario básico proporcional":
                datos_pila[emp_id]["ibc"] += d.valor
            elif d.concepto == "Salud empleado":
                datos_pila[emp_id]["salud_emp"] += d.valor
            elif d.concepto == "Pensión empleado":
                datos_pila[emp_id]["pension_emp"] += d.valor
            elif d.concepto == "Salud empleador":
                datos_pila[emp_id]["salud_pat"] += d.valor
            elif d.concepto == "Pensión empleador":
                datos_pila[emp_id]["pension_pat"] += d.valor
            elif d.concepto == "ARL":
                datos_pila[emp_id]["arl"] += d.valor
            elif d.concepto == "Caja compensación":
                datos_pila[emp_id]["caja"] += d.valor
            elif d.concepto == "SENA":
                datos_pila[emp_id]["sena"] += d.valor
            elif d.concepto == "ICBF":
                datos_pila[emp_id]["icbf"] += d.valor

    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"PILA {periodo}"

    # Encabezados
    headers = [
        "Nombre Empleado", "Documento", "Salario Básico", "IBC",
        "Salud (4%)", "Pensión (4%)", "Salud (8.5%)", "Pensión (12%)",
        "ARL", "Caja (4%)", "SENA (2%)", "ICBF (3%)", "Total Aportes"
    ]
    
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row, emp_data in enumerate(datos_pila.values(), 2):
        total_aportes = (
            emp_data["salud_emp"] + emp_data["pension_emp"] +
            emp_data["salud_pat"] + emp_data["pension_pat"] +
            emp_data["arl"] + emp_data["caja"] +
            emp_data["sena"] + emp_data["icbf"]
        )
        
        ws.cell(row=row, column=1, value=emp_data["nombre"])
        ws.cell(row=row, column=2, value=emp_data["documento"])
        ws.cell(row=row, column=3, value=float(emp_data["salario_base"]))
        ws.cell(row=row, column=4, value=float(emp_data["ibc"]))
        ws.cell(row=row, column=5, value=float(emp_data["salud_emp"]))
        ws.cell(row=row, column=6, value=float(emp_data["pension_emp"]))
        ws.cell(row=row, column=7, value=float(emp_data["salud_pat"]))
        ws.cell(row=row, column=8, value=float(emp_data["pension_pat"]))
        ws.cell(row=row, column=9, value=float(emp_data["arl"]))
        ws.cell(row=row, column=10, value=float(emp_data["caja"]))
        ws.cell(row=row, column=11, value=float(emp_data["sena"]))
        ws.cell(row=row, column=12, value=float(emp_data["icbf"]))
        ws.cell(row=row, column=13, value=float(total_aportes))

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        first_cell = column_cells[0]
        if first_cell.value:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[first_cell.column_letter].width = length + 2

    # Guardar en buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def obtener_resumen_mes(db: AsyncSession, year: int, month: int) -> dict:
    """Obtiene resumen de costos de nómina para un mes."""
    fecha_inicio = date(year, month, 1)
    if month == 12:
        fecha_fin = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        fecha_fin = date(year, month + 1, 1) - timedelta(days=1)

    query = select(Nomina).where(
        Nomina.fecha_inicio >= fecha_inicio,
        Nomina.fecha_fin <= fecha_fin,
        Nomina.estado.in_(["APROBADO", "PAGADO"])
    )
    result = await db.execute(query)
    nominas = result.scalars().all()

    total_neto = sum((n.neto_pagar for n in nominas), Decimal("0"))
    total_devengado = sum((n.total_devengado for n in nominas), Decimal("0"))
    total_deducciones = sum((n.total_deducciones for n in nominas), Decimal("0"))
    count = len(nominas)

    return {
        "periodo": f"{year}-{month:02d}",
        "nominas_liquidadas": count,
        "total_devengado": total_devengado,
        "total_deducciones": total_deducciones,
        "total_neto_pagar": total_neto,
    }


async def obtener_reporte_horas_extra(db: AsyncSession, fecha_inicio: date, fecha_fin: date) -> list[dict]:
    """Obtiene detalle de horas extra por empleado en un rango."""
    query = (
        select(DetalleNomina, Nomina)
        .join(Nomina)
        .where(
            Nomina.fecha_inicio >= fecha_inicio,
            Nomina.fecha_fin <= fecha_fin,
            DetalleNomina.categoria == "DEVENGADO",
            DetalleNomina.concepto.like("%Horas extra%"),
            DetalleNomina.horas > 0
        )
        .options(selectinload(Nomina.empleado))
    )
    result = await db.execute(query)
    rows = result.all()

    reporte = []
    for detalle, nomina in rows:
        reporte.append({
            "empleado": nomina.empleado.nombre,
            "periodo": f"{nomina.fecha_inicio} a {nomina.fecha_fin}",
            "tipo_extra": detalle.concepto,
            "horas": float(detalle.horas) if detalle.horas else 0,
            "valor": float(detalle.valor)
        })

    return reporte
